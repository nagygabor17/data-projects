from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, colors, Alignment

workbook = load_workbook("pelda.xlsx")
active_sheet = workbook.active
print(active_sheet["B1"])
print(active_sheet["B1"].value)

print(active_sheet.cell(row=2, column=2)) #1-től indexel az openpyxl
print(active_sheet.cell(row=2, column=2).value)

active_sheet["B11"].value = "Szia"

for x in active_sheet["A"]:
    if x.value is None: #Cella mindig lesz, de az értéke lesz None, ha üres
        continue
    print(x)

    x.value = x.value + "HAHA"
    #Lehet stílust is beállítani
    x.font = Font(bold=True, italic=True, color=colors.BLUE)

workbook.save("masik.xlsx")


#Természetesen arra is van lehetőség, hogy új excel fájlt készítsünk megnyitás nélkül

uj_munkafuzet = Workbook() #Be kell importálni
aktiv_munkalap = uj_munkafuzet.active
aktiv_munkalap.title = "Átnevezett Munkalap"

masik_munkalap = uj_munkafuzet.create_sheet("Ez lesz az elso", 0) # A második paraméter pozicionál a munkalapok között
aktiv_munkalap["A1"] = "Fejlec 1"
aktiv_munkalap["B1"] = "Fejlec 2"

masik_munkalap["A1"] = "Ez is fejléc"

uj_munkafuzet.save(filename = "ujexcel.xlsx")