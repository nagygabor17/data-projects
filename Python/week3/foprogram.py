import os
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
dirname ="./inputs"
input_files = os.listdir(dirname)
print(input_files)
data = []
common_header = None
for f in input_files:
    if not f.endswith(".csv"):
        continue
    with open(f"{dirname}/{f}", "r") as inputfile:
        csv_reader = csv.reader(inputfile, delimiter=',')
        current_header = next(csv_reader)
        if common_header is None:
            common_header = current_header
        else:
            if current_header != common_header:
                raise ValueError("Nem egyforma formátum")
        for row in csv_reader:
            if row in data: #Duplikáció kiszűrése
                continue
            data.append(row)
print(data)

now = datetime.now()
formatted_time = now.strftime("%Y%m%d%H%M%S")
print(formatted_time)

wb = Workbook()
wb.active.title="Összefűzve"

wb.active.append(common_header)
for d in data:
    wb.active.append(d)
for x in range(1,5):
    wb.active.cell(1, x).font = Font(bold=True, italic=True)
    wb.active.cell(1, x).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    wb.active.cell(1, x).alignment = Alignment(horizontal="center")

wb.save(f"osszefuzott_{formatted_time}.xlsx")


